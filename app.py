import os
import sys
import glob
from datetime import datetime
import pandas as pd
from tkinter import *
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import subprocess
import ttkbootstrap as tb
from tkinter import simpledialog
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

STUDENTS_DIR = "students"
os.makedirs(STUDENTS_DIR, exist_ok=True)

DEFAULT_COLUMNS = ["class no.", "date", "day", "class timing", "comment"]

def default_columns(path, sheet_name = "Sheet1"):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    for i in range(len(headers) - 1,-1, -1):
        if (headers[i] == None):
            headers.pop()
        else:
            break

    return headers

def list_student_names():
    files = glob.glob(os.path.join(STUDENTS_DIR, "*.xlsx"))
    return sorted([os.path.splitext(os.path.basename(f))[0] for f in files])

def normalize_headers(path):
    wb = load_workbook(path)
    ws = wb.active

    headers = default_columns(path)

    cleaned = [str(h).strip().lower() if h is not None else "" for h in headers]
    green_fill = PatternFill(start_color = "92d050", end_color = "92d050", fill_type = "solid")
    
    for col_index, header in enumerate(cleaned, start = 1):
        ws.cell(row = 1, column=col_index).value = header
        ws.cell(row = 1, column=col_index).fill = green_fill
    
    ws.title = "Sheet1"
    wb.save(path)
    
def ensure_student_file(name):
    path = os.path.join(STUDENTS_DIR, f"{name}.xlsx")

    if not os.path.exists(path):
        df = pd.DataFrame(columns = DEFAULT_COLUMNS)
        df.to_excel(path, index = False)
    
    normalize_headers(path)
    return path

def get_next_class_number(path,sheet_name="Sheet1",column_name = "class no."):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    headers = default_columns(path)
    
    try:
        col_idx = headers.index(column_name) + 1
    except ValueError:
        raise Exception(f"Column '{column_name}' not found")

    for row in range(ws.max_row,1,-1):
        value = ws.cell(row = row,column = col_idx).value
        if value not in (None, ""):
            return int(value) + 1
    
    return 1
   

def append_attendance_to_excel(name, date_str, time_str, comment = ""):
    path = ensure_student_file(name)
    wb = load_workbook(path)
    ws = wb.active

    headers = default_columns(path)
    
    try:
        d = datetime.strptime(date_str, "%d-%m-%Y")
        day = d.strftime("%A")
        date_fmt = d.strftime("%d-%m-%Y")
    except Exception:
        day = ""
        date_fmt = date_str

    #Compute next class number
    if comment in ["Attended", "Absent"]:
        next_class_no = get_next_class_number(path)
    else:
        next_class_no = ""
    
    new_row = {
        "class no.": next_class_no,
        "date":date_fmt,
        "day": day,
        "class timing": time_str,
        "comment": comment
    }

    ordered_row = [new_row.get(col.strip().lower(),"") for col in headers]
    ws.append(ordered_row)


    last_row = ws.max_row
    blue_fill = PatternFill(start_color="92cddc", end_color="92cddc", fill_type="solid")
    skin_fill = PatternFill(start_color="e6b8b7", end_color="e6b8b7", fill_type="solid")

    comment_col_idx = headers.index("comment") + 1

    if comment in ["Attended", "Absent"]:
        ws.cell(row = last_row,column = comment_col_idx).fill = blue_fill
        ws.cell(row = last_row,column = comment_col_idx).font = Font(bold = True)
    
    else:
        for col in range(1, len(headers) + 1):
            if col == len(headers):
                ws.cell(row = last_row, column = col).font = Font(bold = True)
            ws.cell(row = last_row, column = col).fill = skin_fill
    wb.save(path)


class AutocompleteCombobox(ttk.Combobox):
    def __init__(self, master = None, **kwargs):
        super().__init__(master, **kwargs)
        self._all_values = []
        self.bind('<KeyRelease>', self._on_keyrelease)

    def set_values(self, values):
        self._all_values = list(values)
        self["values"] = self._all_values

    def _on_keyrelease(self, event):
        typed = self.get().strip().lower()
        if typed == "":
            filtered = self._all_values
        else:
            filtered = [v for v in self._all_values if typed in v.lower()]
            if "Add New Student" not in filtered:
                filtered.append("Add New Student")
        self["values"] = filtered
        if filtered:
            self.event_generate('<Down>')

    
class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("White Dove Foundation")
        self.root.geometry("1000x1000")
        self._build_ui()
        self.refresh_student_list()

    def _build_ui(self):
        pad = 8
        frame = ttk.Frame(self.root, padding = pad)
        frame.pack(fill = BOTH, expand = True)


        row1 = ttk.Frame(frame)
        row1.pack(fill = X, pady = 4)
        ttk.Label(row1, text = "Student (type):").pack(side = LEFT, padx = (0, 6))

        self.combo = AutocompleteCombobox(row1, width = 40)
        self.combo.pack(side = LEFT, padx = (0, 6))
        ttk.Button(row1, text = "Refresh", command = self.refresh_student_list).pack(side = LEFT)
        ttk.Button(row1, text = "Add->", command = self.add_selected_student).pack(side = LEFT, padx = (6, 0))

        row2 = ttk.Frame(frame)
        row2.pack(fill = BOTH, pady = 8, expand = True)


        left = ttk.Frame(row2)
        left.pack(side = LEFT, fill = BOTH, expand = True, padx = (0,6))
        ttk.Label(left, text = "Students to mark:").pack(anchor = W)
        self.listbox = Listbox(left, height = 12)
        self.listbox.pack(fill = BOTH, expand = True)

        btns = ttk.Frame(row2)
        btns.pack(side = LEFT, fill = Y)
        ttk.Button(btns, text = "Remove selected", command = self.remove_selected).pack(pady = 6)
        ttk.Button(btns, text = "Clear List", command = self.clear_list).pack(pady = 6)

        right = ttk.Frame(frame)
        right.pack(fill = X, pady = (6, 0))

        #Date
        date_row = ttk.Frame(right)
        date_row.pack(fill = X, pady = 6)
        ttk.Label(date_row, text = "Date (DD-MM-YYYY):").pack(side = LEFT, padx = (0, 6))

        self.date_picker = tb.DateEntry(date_row,dateformat='%d-%m-%Y',bootstyle="danger",firstweekday = 0)
        self.date_picker.pack(side = LEFT)

        #Start Time
        start_time_row = ttk.Frame(right)
        start_time_row.pack(fill= X, pady = 6)
        ttk.Label(start_time_row, text = "Class Start Time (HH:MM)").pack(side = LEFT, padx = (0, 6))

        self.start_hour_var = StringVar(value = "12")
        start_hour_spin = ttk.Spinbox(start_time_row, from_ = 0, to = 12, wrap = True,textvariable=self.start_hour_var, width = 5, command = self.update_time_var, state="readonly")
        start_hour_spin.pack(side = LEFT)

        self.start_min_var = StringVar(value = "00")
        start_min_spin = ttk.Spinbox(start_time_row, from_= 0, to=59, wrap = True, textvariable=self.start_min_var, width = 5, format = "%02.0f", command = self.update_time_var, state="readonly")
        start_min_spin.pack(side = LEFT)

        self.start_ampm = StringVar(value = "AM")
        start_ampm_spin = ttk.Spinbox(start_time_row, values = ("AM", "PM"), wrap = True, textvariable=self.start_ampm,width = 5, command = self.update_time_var,state = "readonly")
        start_ampm_spin.pack(side = LEFT)

        #End Time
        end_time_row = ttk.Frame(right)
        end_time_row.pack(fill= X, pady = 6)
        ttk.Label(end_time_row, text = "End Start Time (HH:MM)").pack(side = LEFT, padx = (0, 6))

        self.end_hour_var = StringVar(value = "12")
        end_hour_spin = ttk.Spinbox(end_time_row, from_ = 0, to = 12, wrap = True,textvariable=self.end_hour_var, width = 5, command=self.update_time_var, state="readonly")
        end_hour_spin.pack(side = LEFT)

        self.end_min_var = StringVar(value = "00")
        end_min_spin = ttk.Spinbox(end_time_row, from_= 0, to=59, wrap = True, textvariable=self.end_min_var, width = 5, format = "%02.0f",command= self.update_time_var, state="readonly")
        end_min_spin.pack(side = LEFT)

        self.end_ampm = StringVar(value = "AM")
        end_ampm_spin = ttk.Spinbox(end_time_row, values = ("AM", "PM"), wrap = True, textvariable=self.end_ampm,width = 5, command = self.update_time_var,state = "readonly")
        end_ampm_spin.pack(side = LEFT)

        self.time_var = StringVar()
        self.update_time_var()


        # Comment dropdown
        remark_row = ttk.Frame(right)
        remark_row.pack(fill = X, pady = 6)
        ttk.Label(remark_row, text = "Comment:").pack(side = LEFT, padx = (0, 6))
        self.remark_combo = ttk.Combobox(
            remark_row,
            values = ["Attended", "Absent", "Class Not Counted", "Exam Leave", "Holiday Leave", "Free Class"],
            state = "readonly",
            width = 20
        )
        self.remark_combo.bind("<<ComboboxSelected>>", self.update_comment_var)
        self.remark_combo.current(0)
        self.remark_combo.pack(side = LEFT)

        self.comment_var = StringVar()
        self.update_comment_var()

        #Save button
        mark_row = ttk.Frame(right)
        mark_row.pack(fill = X, pady = 12)
        ttk.Button(mark_row, text = "Mark for listed students",command = self.mark_for_list).pack()

        #Log
        status_row2 = ttk.Frame(frame)
        status_row2.pack(fill = BOTH, pady = (8, 0), expand = True)
        ttk.Label(status_row2, text = "Log:").pack(anchor = W)
        self.log_text = Text(status_row2, height = 7, state = DISABLED)
        self.log_text.pack(fill = BOTH, expand = True)

    def update_comment_var(self, event = None):
        if self.remark_combo.get() == "Holiday Leave":
            holiday_reason = simpledialog.askstring("Holiday Leave", "Enter which holiday this leave is provided for:")
            if not holiday_reason:
                holiday_reason = "Unspecified"
            holiday = f"Holiday Leave: {holiday_reason}"
            self.comment_var.set(holiday)
        
        else:
            self.comment_var.set(self.remark_combo.get())
            
    
    def update_time_var(self):
        selected_time = f"{self.start_hour_var.get()}:{self.start_min_var.get()} {self.start_ampm.get()} to {self.end_hour_var.get()}:{self.end_min_var.get()} {self.end_ampm.get()}"
        self.time_var.set(selected_time)
        
    def log(self, msg):
        self.log_text.configure(state = NORMAL)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert(END, f"[{now}] {msg}\n")
        self.log_text.see(END)
        self.log_text.configure(state = DISABLED)

    def refresh_student_list(self):
        names = list_student_names()

        names.append("Add New Student")

        self.combo.set_values(names)
        self.log(f"Loaded {len(names) - 1} student files.")
    
    def add_selected_student(self):
        name = self.combo.get().strip()

        if not name:
            messagebox.showwarning("No name", "Type or select a student name.")
            return 
        
        if name == "Add New Student":
            new_name = simpledialog.askstring("New Student", "Enter new student name:")
            if not new_name:
                return
            
            new_name = new_name.strip()
            if not new_name:
                return
            

            #If file exists, just add
            path = os.path.join(STUDENTS_DIR, f"{new_name}.xlsx")
            if not os.path.exists(path):
                ensure_student_file(new_name)
                self.log(f"Created file for {new_name}")

            self.refresh_student_list()

            if new_name not in self.listbox.get(0, END):
                self.listbox.insert(END, new_name)
                self.log(f"Added {new_name}")
            
            return
        

        #Normal flow for existing names
        path = os.path.join(STUDENTS_DIR, f"{name}.xlsx")
        if not os.path.exists(path):
            if messagebox.askyesno("Create file?", f"No file for '{name}'. Create?"):
                ensure_student_file(name)
                self.refresh_student_list()
            else:
                return
            
        if name in self.listbox.get(0, END):
            messagebox.showinfo("Already", f"{name} is already in the list.")
            return
        
        self.listbox.insert(END, name)
        self.log(f"Added {name}")

    
    def remove_selected(self):
        sel = self.listbox.curselection()
        for i in reversed(sel):
            name = self.listbox.get(i)
            self.listbox.delete(i)
            self.log(f"Removed {name}")

    
    def clear_list(self):
        self.listbox.delete(0, END)
        self.log("Cleared list.")

    def mark_for_list(self):
        count = self.listbox.size()
        if count == 0:
            messagebox.showinfo("List is empty", "No students to mark")
            return
        
        date_str = self.date_picker.entry.get()
        time_str = self.time_var.get()
        comment = self.comment_var.get().strip()

        if not messagebox.askyesno("Confirm", f"Mark {count} students as '{comment}' on {date_str} {time_str}?"):
            return
        
        names = self.listbox.get(0, END)

        for name in names:
            err_student = []
            try:
                append_attendance_to_excel(name, date_str, time_str, comment)
                self.log(f"Marked {name} as {comment}")

            except Exception as e:
                err_student.append(name + ".xlsx")
                self.log(f"Error writing {name}: {e}")
        
        if err_student:
            messagebox.showerror(
            "Excel files open",
            "Please close the following Excel files:\n\n"
            + "\n".join(err_student) + "\n\nMarked attendance for rest if any. Please try again after closing the file(s)."
            )

        else:
            messagebox.showinfo("Done", "Marked attendance for all listed students.")

def get_open_excel_files(folder):
    open_files = []

    for filename in os.listdir(folder):
        if filename.lower().endswith(".xlsx"):
            path = os.path.join(folder, filename)
            try:
                with open(path, "a"):
                    pass
            except PermissionError:
                if not filename.startswith("~$"):
                    open_files.append(filename)
    
    return open_files

def main():
    open_files = get_open_excel_files(STUDENTS_DIR)

    if open_files:
        messagebox.showerror(
            "Excel files open",
            "Please close the following Excel files and restart the application:\n\n"
            + "\n".join(open_files) + "\n\nPlease don't open any excel files while the application is running."
        )

        return

    root = tb.Window(themename="darkly")
    app = AttendanceApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()

            
            








